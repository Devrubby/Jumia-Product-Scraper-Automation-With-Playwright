from openpyxl.utils import get_column_letter
from openpyxl.styles import  Font, Alignment, Border, Side

BOLD_FONT = Font(bold=True)

THIN_BOX = Border(top=Side(style='thin'), bottom=Side(style='thin'),
                  left=Side(style='thin'), right=Side(style='thin'))

def style_columns(sheet):
    for cell in sheet[1]:
        cell.font = BOLD_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BOX
    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[column_letter].width = max_length + 5


def format_currency_columns(sheet, columns_index=None):
    naira_format = '"₦"#,##0;("₦"#,##0);"-"'
    if columns_index is None:
        columns_index = []
    for row in range(2, sheet.max_row + 1):
        for value in columns_index:
            cell = sheet.cell(row=row, column=value)
            if cell.value is not None:
                cell.number_format = naira_format

def center_align_column(ws, target_column):
    center_alignment = Alignment(horizontal='center', vertical='center')
    for row in range(1, ws.max_row + 1):
        ws.cell(row=row, column=target_column).alignment = center_alignment


def create_products_sheet(workbook, data):
    ws = workbook.active
    ws.title = "Products Info"
    ws.append(list(data))
    for products in list(data.values):
        ws.append(list(products))
    style_columns(ws)
    format_currency_columns(ws, [2, 3])
    ws.column_dimensions["A"].width = 40
    for i in range(2, ws.max_row):
        ws[f"A{i}"].alignment = Alignment(wrap_text=True)
    center_align_column(ws, 4)
    center_align_column(ws, 7)




def create_summary_sheet(workbook, data, search_term):
    ws = workbook.create_sheet("Summary Statistics")
    ws.append(["Metric", "Value"])
    ws.append(["Search Term", search_term])
    for key, value in data.items():
        if isinstance(value, dict):
            continue
        ws.append([key, value])
    style_columns(ws)

def create_insight_sheet(workbook, data):
    ws = workbook.create_sheet("Product Insights")
    row = 1
    column_a_longest_text = 0
    for section_title, product_details in data.items():
        if isinstance(product_details, dict):
            ws.cell(row=row, column=1).value = section_title
            ws.cell(row=row, column=1).font = BOLD_FONT
            column_a_longest_text = max(column_a_longest_text, len(section_title))
            row += 1
            for key, value in product_details.items():
                ws.cell(row=row, column=1).value = key
                ws.cell(row=row, column=2).value = value
                row += 1
    ws.column_dimensions["A"].width = column_a_longest_text + 5
    ws.column_dimensions["B"].width = 40
    for i in range(2, ws.max_row):
        ws[f"B{i}"].alignment = Alignment(wrap_text=True)

def create_brand_sheet(workbook, data):
    ws = workbook.create_sheet("Brand Analysis")
    ws.append(["Brand", "Count"])
    for key, value in data.items():
        ws.append([key, value])
    style_columns(ws)

def create_price_range_sheet(workbook, data):
    ws = workbook.create_sheet("Price Range Distribution")
    ws.append(["Price Range", "Count"])
    for key, value in data.items():
        ws.append([key, value])
    style_columns(ws)